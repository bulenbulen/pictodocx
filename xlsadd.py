import pandas as pd
from openpyxl import load_workbook as lw

#modname函数返回最后一个非空值，modpathname函数返回xx-xx-xx-xx格式的字符串列表
#excel转换成list
#excel文件路径
def modname(filepath,sheetpage):
    modlist = []
    
    wb = lw(filepath, data_only=False)
    ws = wb[sheetpage]

    # 先读原始表格值；随后只把真实合并区域的空白格补成顶端单元格的值
    rows = [list(row) for row in ws.iter_rows(values_only=True)]

    for merged_range in ws.merged_cells.ranges:
        start_row = merged_range.min_row
        start_col = merged_range.min_col
        end_row = merged_range.max_row
        end_col = merged_range.max_col
        value = ws.cell(start_row, start_col).value

        for r in range(start_row, end_row + 1):
            for c in range(start_col, end_col + 1):
                rows[r - 1][c - 1] = value

    wb.close()

    df = pd.DataFrame(rows)

    # 第一行是表头，不需要输出
    for _, targetrow in df.iloc[1:].iterrows():
        valid_values = [item for item in targetrow.tolist() if pd.notna(item) and str(item).strip() != '']
        if not valid_values:
            continue
#最后一个非空值
        last_value = valid_values[-1]
        modlist.append(last_value)

    return modlist


def modpathname(filepath,sheetpage):
    
    joinlist = []
    wb = lw(filepath, data_only=False)
    ws = wb[sheetpage]

    # 先读原始表格值；随后只把真实合并区域的空白格补成顶端单元格的值
    rows = [list(row) for row in ws.iter_rows(values_only=True)]

    for merged_range in ws.merged_cells.ranges:
        start_row = merged_range.min_row
        start_col = merged_range.min_col
        end_row = merged_range.max_row
        end_col = merged_range.max_col
        value = ws.cell(start_row, start_col).value

        for r in range(start_row, end_row + 1):
            for c in range(start_col, end_col + 1):
                rows[r - 1][c - 1] = value

    wb.close()

    df = pd.DataFrame(rows)

    # 第一行是表头，不需要输出
    for _, targetrow in df.iloc[1:].iterrows():
        valid_values = [item for item in targetrow.tolist() if pd.notna(item) and str(item).strip() != '']
        if not valid_values:
            continue
#xx-xx-xx-xx
        join_values = [str(item).strip() for item in valid_values]
        joinlist.append('-'.join(join_values))

    return joinlist











#result, join_result = xlsadd(r'D:/DATA/doctest/list.xlsx')
#print('当前显示:', result)
#print('xx-xx-xx-xx:', join_result)


#print(df)

#print(f"原始数据形状: {df.shape}")          # 输出 (行数, 列数)
#print(f"有效数据行数: {len(df)}")


#指定行
#targetrow = df.iloc[2]  
#print(list(targetrow))
#xx-xx-xx-xx格式 ，报错就加转换成字符串dropna().astype(str)，dropna()去掉空值，astype(str)转换成字符串
#funmodname= '-'.join(targetrow.dropna().astype(str))
#最后一个非空值
#lastmodname = targetrow.dropna().iloc[-1]
#print(funmodname,lastmodname)
